import openpyxl as xl
import wx
import datetime
import os
import shutil
import getpass as gt
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import re

setores = ['IE-INJECAO', 'IE-PINTURA', 'IE-MONTAGEM']
opcoes  = ['SIM', 'NAO']

def leitura():
    usuario   = gt.getuser()
    continuar = 'SIM'
    gerada    = f"C:\\Users\\{usuario}\\Desktop\\REPORT_MASTER.xlsx"
    original  = r'\\files-gdbr01\\GDBR\\ADMINISTRATION\\IT\\Desenvolvimento\\Apontamentos de producao - Programas\\Um-a-Um_Homologado\\modelo\\REPORT_Master__FUNCIONAL.xlsx'
    target    = gerada

    shutil.copyfile(original, target)
    wb_modelo    = xl.load_workbook(gerada, data_only=False)
    ws_modeloPec = wb_modelo['REPORT']

    cont          = 2
    item          = []
    total         = []
    maquina       = []
    ng            = []
    usuario       = []
    borra         = []
    kanbanModelo  = []
    partNumber    = []
    indice        = 0

    dlg = wx.TextEntryDialog(None, 'Informe o dia: XX/XX/XXXX','Dialog')  
    if dlg.ShowModal() == wx.ID_OK:
        data = str(dlg.GetValue())
        
        while True:
            if dlg.ShowModal() == wx.ID_OK:
                data = str(dlg.GetValue())
                if re.match(r'^\d{2}/\d{2}/\d{4}$', data):
                    try:
                        datetime.datetime.strptime(data, "%d/%m/%Y")
                        break  
                    except ValueError:
                        wx.MessageBox("Data inválida! Use o formato DD/MM/AAAA.", "Erro", wx.OK | wx.ICON_ERROR)
                else:
                    wx.MessageBox("Formato de data inválido! Use DD/MM/AAAA.", "Erro", wx.OK | wx.ICON_ERROR)
            else:
                dlg.Destroy()
                return
        dlg.Destroy()
    

    while continuar == 'SIM':

        with wx.FileDialog(None, "Selecione arquivos Excel", wildcard="Excel Files (*.xlsm;*.xlsx)|*.xlsm;*.xlsx", style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE) as fileDialog:
            
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            
            pathnames = fileDialog.GetPaths()
            print(f"\nArquivos selecionados: {pathnames}")

        dlg = wx.SingleChoiceDialog(None, ("Escolha seu setor"), "SETOR", setores)

        if dlg.ShowModal() == wx.ID_OK:
            setor = dlg.GetStringSelection()
            print(f"\nSetor: {setor}")
            setores.remove(setor)
        dlg.Destroy()

        for pathname in pathnames:

            print(f"\nProcessando: {pathname}")

            wb_reporte = xl.load_workbook(pathname, data_only=True)
            ws_reporte = wb_reporte['Cadastro']

            maxrow = ws_reporte.max_row
    
            for i in range (3, maxrow+1):
                if ws_reporte.cell(row=i, column=3).value is not None and  ws_reporte.cell(row=i, column=4).value is not None:
                    kanban = ws_reporte.cell(row=i, column=3).value #Pega valores da coluna C da aba cadastro (Valor do Kanban)
                    kanbanModelo.append(kanban)

                    pn = ws_reporte.cell(row=i, column=4).value #Pega valores da coluna D da aba cadastro (Part Number)
                    partNumber.append(pn)

            ws_reporte = wb_reporte['Apontamentos']

            maxcol = ws_reporte.max_column
            maxrow = ws_reporte.max_row

            # i --> linha
            # j --> coluna
            for i in range(12, maxrow+1):
                val3 = ws_reporte.cell(row=i, column=3).value #Coluna3

                # Verifica se as colunas 1 e 2 possuem valores
                if val3 is not None and val3 not in (None, ""):

                    # Obtém o valor da célula
                    item.append(ws_reporte.cell(row=i, column=3).value)
                    if item[indice] in kanbanModelo:
                        idx = kanbanModelo.index(item[indice])
                        ws_modeloPec.cell(row=cont, column=2).value = partNumber[idx]

                    total.append(ws_reporte.cell(row=i, column=7).value)
                    maquina.append(ws_reporte.cell(row=i, column=6).value)
                    ng.append(ws_reporte.cell(row=i, column=65).value)
                    usuario.append(ws_reporte.cell(row=i, column=2).value)
                    borra_val = ws_reporte.cell(row=i, column=68).value
                    if borra_val is None or borra_val == "":
                        borra_val = 0
                    borra.append(borra_val)

                # 

                    for j in range(1, maxcol+1):

                        if not isinstance(ws_modeloPec.cell(row=cont, column=j), MergedCell):
                            ws_modeloPec.cell(row=cont, column=1).value = item[indice]
                            ws_modeloPec.cell(row=cont, column=3).value = total[indice]
                            ws_modeloPec.cell(row=cont, column=23).value = maquina[indice]
                            ws_modeloPec.cell(row=cont, column=4).value = ng[indice]
                            ws_modeloPec.cell(row=cont, column=21).value = usuario[indice]
                            ws_modeloPec.cell(row=cont, column=22).value = data
                            ws_modeloPec.cell(row=cont, column=20).value = borra[indice]
                            ws_modeloPec.cell(row=cont, column=25).value = setor
                            ws_modeloPec.cell(row=cont, column=j).alignment = Alignment(horizontal='center')
                            
                    indice += 1
                    cont += 1

        wb_modelo.save(gerada)

        dlg = wx.SingleChoiceDialog(None, ("Deseja continuar?"), "CONTINUAR", opcoes)

        if dlg.ShowModal() == wx.ID_OK:
            continuar = dlg.GetStringSelection()
        dlg.Destroy()


